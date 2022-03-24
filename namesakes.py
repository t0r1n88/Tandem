"""
Скрипт для нахождения однофамильцев и полных тезок в общей таблице

"""

import pandas as pd
import os
# from docxtpl import DocxTemplate
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import time
import datetime
from datetime import date
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import BarChart, Reference, PieChart, PieChart3D, Series
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
import sys
import locale

wb = openpyxl.Workbook()
wb.create_sheet(title='Однофамильцы',index=0)
wb.create_sheet(title='Фамилия + Имя',index=1)
wb.create_sheet(title='Полные тезки',index=2)

general_table = 'Общий список БРИТ 26.11.2021.xlsx'

path_to_end_folder = 'resources/'

df = pd.read_excel(general_table)


# Ищем однофамильцев
dupl_fam_df = df[df.duplicated(['Фамилия'],keep=False)]
# Сортируем
sort_dupl_fam_df = dupl_fam_df.sort_values(by='Фамилия')

for r in dataframe_to_rows(sort_dupl_fam_df, index=True, header=True):
    wb['Однофамильцы'].append(r)
wb['Однофамильцы'].column_dimensions['B'].width = 30

# Ищем Фамилия +Имя

dupl_fam_name_df = df[df.duplicated(['Фамилия','Имя'],keep=False)]
sort_dupl_fam_name_df = dupl_fam_name_df.sort_values(by='Фамилия')

for r in dataframe_to_rows(sort_dupl_fam_name_df, index=True, header=True):
    wb['Фамилия + Имя'].append(r)
wb['Фамилия + Имя'].column_dimensions['B'].width = 30


# полные тезки
namesakes_df =  df[df.duplicated(['Фамилия','Имя','Отчество'],keep=False)]
sort_namesakes_df = namesakes_df.sort_values(by='Фамилия')

for r in dataframe_to_rows(sort_namesakes_df, index=True, header=True):
    wb['Полные тезки'].append(r)
wb['Полные тезки'].column_dimensions['B'].width = 30
t = time.localtime()
current_time = time.strftime('%H_%M_%S', t)
# Сохраняем итоговый файл
wb.save(f'{path_to_end_folder}/Однофамильцы,Полные тезки от {current_time}.xlsx')
